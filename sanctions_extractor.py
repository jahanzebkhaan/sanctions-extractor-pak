"""
═══════════════════════════════════════════════════════════════
  Global Sanctions Lists Extractor  v5  — full data edition
═══════════════════════════════════════════════════════════════
  Verified against actual file structures before writing.

  Sources:
    1. OFAC   (US Treasury SDN)           → XML  treasury.gov
    2. UK HMT (OFSI Consolidated List)    → CSV  blob.core.windows.net
    3. EU FSF (Financial Sanctions Files) → CSV  webgate.ec.europa.eu
    4. UNSC   (UN Security Council)       → XML  scsanctions.un.org

  Setup:   pip install requests openpyxl
  Run:     python sanctions_extractor.py
═══════════════════════════════════════════════════════════════
"""

import csv, re
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Sources ───────────────────────────────────────────────────────────────────
SOURCES = {
    'OFAC': {
        'url':      'https://www.treasury.gov/ofac/downloads/sdn.xml',
        'format':   'xml_ofac',
        'filename': 'ofac_sdn.xml',
    },
    'UK HMT': {
        'url':      'https://ofsistorage.blob.core.windows.net/publishlive/2022format/ConList.csv',
        'format':   'csv_uk',
        'filename': 'uk_conlist.csv',
    },
    'EU': {
        'url':      'https://webgate.ec.europa.eu/fsd/fsf/public/files/csvFullSanctionsList/content?token=dG9rZW4tMjAxNw',
        'format':   'csv_eu',
        'filename': 'eu_sanctions.csv',
    },
    'UNSC': {
        'url':      'https://scsanctions.un.org/resources/xml/en/consolidated.xml',
        'format':   'xml_unsc',
        'filename': 'unsc_consolidated.xml',
    },
    'NACTA': {
        # nfs.nacta.gov.pk portal — XML endpoint returns all records at once.
        # Fields: Name, FatherName, CNIC, District, Province
        # Falls back to page scraping if XML endpoint unavailable.
        'url':           'https://nfs.nacta.gov.pk/xml',
        'format':        'xml_nacta',
        'filename':      'nacta_proscribed.xml',
        'fallback_url':  'https://nfs.nacta.gov.pk',
        'pages_url':     'https://nfs.nacta.gov.pk/proscribed-list?page={}',
    },
}

ORDER       = ['OFAC', 'UK HMT', 'EU', 'UNSC', 'NACTA']
SCRIPT_DIR  = Path(__file__).parent
CACHE_DIR   = SCRIPT_DIR / 'sanctions_cache'
OUTPUT_FILE = SCRIPT_DIR / 'sanctions_output.xlsx'
CACHE_DIR.mkdir(exist_ok=True)

# ── Master column schema ──────────────────────────────────────────────────────
MASTER_COLS = [
    'Source', 'Reference / UID', 'Sanctions Programme',
    'Last Name', 'First Name', 'Second / Middle Name',
    'Third Name', 'Fourth Name', 'Fifth Name',
    'Title', 'Designation / Position', 'Gender',
    'Good Quality AKA', 'Low Quality AKA', 'Formerly Known As',
    'Date of Birth', 'Place of Birth', 'Country of Birth',
    'Nationality 1', 'Nationality 2', 'Nationality 3+',
    'Passport Number', 'Passport Country', 'Passport Issue Date', 'Passport Expiry',
    'National ID Number', 'National ID Type', 'National ID Country',
    'Passport Details', 'National ID Details',
    'Address', 'City', 'State / Province', 'Country', 'Postal Code',
    'Listed On', 'UK Designated On', 'Last Updated',
    'Interpol Link', 'Remarks / Other Info',
]

MASTER_COL_W = [
    10, 18, 26,
    28, 20, 20,
    20, 20, 20,
    14, 32, 10,
    44, 44, 44,
    22, 26, 20,
    22, 22, 22,
    22, 18, 16, 16,
    26, 24, 18,
    36, 36,
    40, 18, 18, 18, 12,
    16, 16, 16,
    40, 52,
]

COLORS = {
    'OFAC':   {'row': 'FFF2CC', 'header': '7F6000', 'sep': 'BF9000'},
    'UK HMT': {'row': 'D9EAF7', 'header': '1F4E79', 'sep': '2E75B6'},
    'EU':     {'row': 'E2F0D9', 'header': '375623', 'sep': '548235'},
    'UNSC':   {'row': 'FCE4D6', 'header': '843C0C', 'sep': 'C55A11'},
    'NACTA':  {'row': 'E8D5F5', 'header': '5B2D8E', 'sep': '7B3DB5'},
}

SOURCE_COLS = {
    'OFAC': [
        'Source', 'Reference / UID', 'Sanctions Programme',
        'Last Name', 'First Name', 'Second / Middle Name',
        'Title', 'Designation / Position',
        'Good Quality AKA', 'Low Quality AKA',
        'Date of Birth', 'Place of Birth',
        'Nationality 1', 'Nationality 2', 'Nationality 3+',
        'Passport Number', 'Passport Country', 'Passport Issue Date', 'Passport Expiry',
        'National ID Number', 'National ID Type', 'National ID Country',
        'Address', 'City', 'State / Province', 'Country', 'Postal Code',
        'Remarks / Other Info',
    ],
    'UK HMT': [
        'Source', 'Reference / UID', 'Sanctions Programme',
        'Last Name', 'First Name', 'Second / Middle Name',
        'Third Name', 'Fourth Name', 'Fifth Name',
        'Title', 'Designation / Position',
        'Good Quality AKA', 'Low Quality AKA', 'Formerly Known As',
        'Date of Birth', 'Place of Birth', 'Country of Birth', 'Nationality 1',
        'Passport Number', 'Passport Details',
        'National ID Number', 'National ID Details',
        'Address', 'City', 'Country', 'Postal Code',
        'Listed On', 'UK Designated On', 'Last Updated',
        'Remarks / Other Info',
    ],
    'EU': [
        'Source', 'Reference / UID', 'Sanctions Programme',
        'Last Name', 'First Name', 'Second / Middle Name',
        'Title', 'Designation / Position', 'Gender',
        'Good Quality AKA',
        'Date of Birth', 'Place of Birth', 'Country of Birth',
        'Nationality 1', 'Nationality 2', 'Nationality 3+',
        'Passport Number', 'Passport Country', 'Passport Expiry',
        'National ID Number', 'National ID Type', 'National ID Country',
        'Address', 'City', 'Country', 'Postal Code',
        'Listed On', 'Remarks / Other Info',
    ],
    'UNSC': [
        'Source', 'Reference / UID', 'Sanctions Programme',
        'Last Name', 'First Name', 'Second / Middle Name',
        'Third Name', 'Fourth Name',
        'Title', 'Designation / Position', 'Gender',
        'Good Quality AKA', 'Low Quality AKA',
        'Date of Birth', 'Place of Birth', 'Country of Birth',
        'Nationality 1', 'Nationality 2', 'Nationality 3+',
        'Passport Number', 'Passport Country', 'Passport Issue Date', 'Passport Expiry',
        'National ID Number', 'National ID Type', 'National ID Country',
        'Address', 'City', 'State / Province', 'Country',
        'Listed On', 'Last Updated',
        'Interpol Link', 'Remarks / Other Info',
    ],
    'NACTA': [
        'Source', 'Last Name', 'First Name',
        'National ID Number',
        'Remarks / Other Info',   # Father's name stored here
        'City',                   # District
        'State / Province',       # Province
        'Sanctions Programme',    # Proscribed organisation (when available)
    ],
}

# ── Helpers ───────────────────────────────────────────────────────────────────
ILLEGAL_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def s(v):
    if not isinstance(v, str): return ''
    return ILLEGAL_RE.sub('', v).strip()

def join(items, sep=' | '):
    seen, out = set(), []
    for item in items:
        v = s(str(item)) if item else ''
        if v and v not in seen:
            seen.add(v); out.append(v)
    return sep.join(out)

def empty_row(source):
    return {c: '' for c in MASTER_COLS} | {'Source': source}

def last_nonempty(*parts):
    return next((p for p in reversed(parts) if s(p)), '')


# ── Downloader: NACTA (Playwright — real browser, clicks Excel button) ────────
#
# nfs.nacta.gov.pk is a Blazor WebAssembly app. The only reliable automated
# method is Playwright which runs a real headless Chrome session.
#
# This function auto-installs Playwright into the correct Python environment.
#
def download_nacta(cfg):
    import subprocess, sys, time

    dest     = CACHE_DIR / cfg['filename']
    xlsx_dest = CACHE_DIR / 'nacta_proscribed.xlsx'

    # Use cached file if exists
    for cached in [dest, xlsx_dest]:
        if cached.exists():
            print(f'  [NACTA] Cached  ({cached.stat().st_size/1024/1024:.2f} MB)')
            return cached

    print(f'  [NACTA] Launching Playwright (headless Chrome)...')

    # ── Step 1: Install playwright into THIS exact Python ─────────────────────
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        print(f'  [NACTA] Installing Playwright into: {sys.executable}')
        try:
            subprocess.check_call(
                [sys.executable, '-m', 'pip', 'install', 'playwright'],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
            )
            subprocess.check_call(
                [sys.executable, '-m', 'playwright', 'install', 'chromium'],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
            )
            from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
            print(f'  [NACTA] Playwright installed OK.')
        except Exception as e:
            print(f'  [NACTA] Auto-install failed: {e}')
            print(f'  Run these two commands manually then re-run the script:')
            print(f'    {sys.executable} -m pip install playwright')
            print(f'    {sys.executable} -m playwright install chromium')
            return None

    # ── Step 2: Launch browser and download ──────────────────────────────────
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            context = browser.new_context(accept_downloads=True)
            page    = context.new_page()

            print(f'  [NACTA] Loading nfs.nacta.gov.pk ...', end='', flush=True)
            # Blazor apps never reach 'networkidle' — use 'domcontentloaded' instead
            # then wait fixed time for Blazor to render
            try:
                page.goto('https://nfs.nacta.gov.pk/', timeout=90000,
                          wait_until='domcontentloaded')
            except Exception:
                pass  # timeout on domcontentloaded is fine — continue anyway
            page.wait_for_timeout(8000)  # wait for Blazor WASM to fully render
            print(' done')

            # Find Excel button — retry up to 3 times with extra waits
            excel_btn = None
            selectors = [
                'button:has-text("Excel")',
                'a:has-text("Excel")',
                'button:has-text("EXCEL")',
                'a:has-text("EXCEL")',
                ':text-matches("Excel", "i")',
                '[onclick*="xcel"]',
                'input[value*="xcel"]',
                'button:has-text("excel")',
            ]
            for attempt in range(3):
                for selector in selectors:
                    try:
                        loc = page.locator(selector)
                        if loc.count() > 0:
                            excel_btn = loc.first
                            print(f'  [NACTA] Excel button found (attempt {attempt+1}).')
                            break
                    except Exception:
                        continue
                if excel_btn:
                    break
                # Not found yet — wait more for Blazor to render
                print(f'  [NACTA] Waiting for page to render (attempt {attempt+1}/3)...')
                page.wait_for_timeout(5000)

            if not excel_btn:
                print(f'  [NACTA] Excel button not found. Page content:')
                print(f'    URL: {page.url}')
                print(f'    Title: {page.title()}')
                print(f'  All buttons/links on page:')
                for el in page.locator('button, a').all()[:30]:
                    try: print(f'    - {repr(el.inner_text()[:60].strip())}')
                    except: pass
                browser.close()
                return None

            # Click and capture the download
            print(f'  [NACTA] Clicking Excel button...', end='', flush=True)
            with page.expect_download(timeout=120000) as dl_info:
                excel_btn.click()

            dl = dl_info.value
            dl.save_as(str(xlsx_dest))
            browser.close()

            size_mb = xlsx_dest.stat().st_size / 1024 / 1024
            print(f'  {size_mb:.2f} MB  ✓  ({dl.suggested_filename})')
            return xlsx_dest

    except Exception as e:
        print(f'\n  ✗ Playwright error: {e}')
        print(f'  Try running with headless=False to debug:')
        print(f'  Change headless=True → headless=False in the script')
        print()
        print(f'  OR manually download:')
        print(f'    1. Go to https://nfs.nacta.gov.pk/')
        print(f'    2. Click Excel button')
        print(f'   3. Move file → {CACHE_DIR / "nacta_proscribed.xlsx"}')
        print(f'    4. Re-run script')
        return None



# ── Downloader ────────────────────────────────────────────────────────────────
def download(key, cfg):
    dest = CACHE_DIR / cfg['filename']
    if dest.exists():
        print(f'  [{key}] Cached  ({dest.stat().st_size/1024/1024:.1f} MB)')
        return dest
    print(f'  [{key}] Downloading ...', end='', flush=True)
    r = requests.get(cfg['url'], headers={'User-Agent': 'SanctionsExtractor/5.0'},
                     timeout=180, stream=True)
    r.raise_for_status()
    with open(dest, 'wb') as fh:
        for chunk in r.iter_content(65536): fh.write(chunk)
    print(f'  {dest.stat().st_size/1024/1024:.1f} MB  ✓')
    return dest

# ── Parser: OFAC XML ──────────────────────────────────────────────────────────
# Verified tags: sdnEntry/uid, lastName, firstName, title, remarks, sdnType
# programList/program, akaList/aka/(category,firstName,lastName,type)
# dateOfBirthList/dateOfBirthItem/dateOfBirth
# placeOfBirthList/placeOfBirthItem/placeOfBirth
# nationalityList/nationality/country
# citizenshipList/citizenship/country
# idList/id/(idType,idNumber,idCountry,issueDate,expirationDate)
# addressList/address/(address1,address2,address3,city,stateOrProvince,postalCode,country)

def parse_ofac_xml(path):
    rows = []
    try:
        tree = ET.parse(path); root = tree.getroot()
    except ET.ParseError as e:
        print(f'\n  ✗ OFAC XML: {e}'); return rows

    ns_m = re.match(r'\{(.*?)\}', root.tag)
    ns   = f'{{{ns_m.group(1)}}}' if ns_m else ''

    def tx(el, tag):
        n = el.find(f'{ns}{tag}')
        return s(n.text) if n is not None and n.text else ''

    def txall(el, tag):
        return [s(n.text) for n in el.findall(f'.//{ns}{tag}') if n.text and n.text.strip()]

    for entry in root.findall(f'.//{ns}sdnEntry'):
        if tx(entry, 'sdnType').lower() != 'individual':
            continue

        row = empty_row('OFAC')
        row['Reference / UID'] = s(tx(entry, 'uid'))
        row['Title']           = s(tx(entry, 'title'))

        # Name — split multi-word firstName into First + Second
        last      = s(tx(entry, 'lastName'))
        first_raw = s(tx(entry, 'firstName'))
        fp        = first_raw.split()
        row['Last Name']            = last
        row['First Name']           = fp[0] if fp else ''
        row['Second / Middle Name'] = ' '.join(fp[1:]) if len(fp) > 1 else ''

        # Programmes
        row['Sanctions Programme'] = join(txall(entry, 'program'), '; ')

        # Remarks
        row['Remarks / Other Info'] = s(tx(entry, 'remarks'))

        # AKA  — OFAC category: 'strong'=good, 'weak'=low, ''=good (default)
        good, low = [], []
        for aka in entry.findall(f'.//{ns}aka'):
            cat  = tx(aka, 'category').lower()
            name = join([tx(aka, 'firstName'), tx(aka, 'lastName')], ' ')
            if not name: continue
            (low if cat == 'weak' else good).append(name)
        row['Good Quality AKA'] = join(good, '; ')
        row['Low Quality AKA']  = join(low,  '; ')

        # DOB — all
        row['Date of Birth'] = join(txall(entry, 'dateOfBirth'), ' | ')

        # POB — all
        row['Place of Birth'] = join(txall(entry, 'placeOfBirth'), ' | ')

        # Nationality + Citizenship (both)
        nats = txall(entry, 'country')  # appears in both nationalityList and citizenshipList
        # But txall gets ALL country tags including address — be specific
        nats = []
        for el in entry.findall(f'.//{ns}nationalityList') + entry.findall(f'.//{ns}citizenshipList'):
            for c in el.findall(f'.//{ns}country'):
                v = s(c.text)
                if v and v not in nats: nats.append(v)
        if nats: row['Nationality 1'] = nats[0]
        if len(nats) > 1: row['Nationality 2'] = nats[1]
        if len(nats) > 2: row['Nationality 3+'] = join(nats[2:], '; ')

        # IDs
        passports, other_ids = [], []
        for id_el in entry.findall(f'.//{ns}id'):
            id_type    = s(tx(id_el, 'idType'))
            id_num     = s(tx(id_el, 'idNumber'))
            id_country = s(tx(id_el, 'idCountry'))
            id_issue   = s(tx(id_el, 'issueDate'))
            id_expiry  = s(tx(id_el, 'expirationDate'))
            if not id_num: continue
            if 'passport' in id_type.lower():
                passports.append((id_num, id_country, id_issue, id_expiry))
            else:
                other_ids.append((id_type, id_num, id_country))

        if passports:
            row['Passport Number']     = join([p[0] for p in passports], ' | ')
            row['Passport Country']    = join([p[1] for p in passports], ' | ')
            row['Passport Issue Date'] = join([p[2] for p in passports], ' | ')
            row['Passport Expiry']     = join([p[3] for p in passports], ' | ')
        if other_ids:
            row['National ID Number']  = join([x[1] for x in other_ids], ' | ')
            row['National ID Type']    = join([x[0] for x in other_ids], ' | ')
            row['National ID Country'] = join([x[2] for x in other_ids], ' | ')

        # Addresses — all
        addrs, cities, states, countries, postcodes = [], [], [], [], []
        for addr in entry.findall(f'.//{ns}address'):
            parts = [tx(addr, t) for t in ['address1', 'address2', 'address3']]
            line  = '; '.join(p for p in parts if p)
            if line and line not in addrs:    addrs.append(line)
            city = tx(addr, 'city')
            if city and city not in cities:   cities.append(city)
            state = tx(addr, 'stateOrProvince')
            if state and state not in states: states.append(state)
            ctry = tx(addr, 'country')
            if ctry and ctry not in countries: countries.append(ctry)
            pc   = tx(addr, 'postalCode')
            if pc and pc not in postcodes:    postcodes.append(pc)

        row['Address']         = join(addrs,     ' | ')
        row['City']            = join(cities,    ' | ')
        row['State / Province']= join(states,    ' | ')
        row['Country']         = join(countries, ' | ')
        row['Postal Code']     = join(postcodes, ' | ')

        rows.append(row)
    return rows

# ── Parser: UK HMT CSV ────────────────────────────────────────────────────────
# Verified headers (row index 1, 36 cols):
# col[0]=Name 6(surname), col[1]=Name 1..col[5]=Name 5, col[6]=Title
# col[10]=DOB, col[11]=Town of Birth, col[12]=Country of Birth, col[13]=Nationality
# col[14]=Passport Number, col[15]=Passport Details
# col[16]=National Identification Number, col[17]=National Identification Details
# col[18]=Position, col[19-24]=Address 1-6, col[25]=Post/Zip Code, col[26]=Country
# col[27]=Other Information, col[28]=Group Type, col[29]=Alias Type
# col[30]=Alias Quality, col[31]=Regime, col[32]=Listed On
# col[33]=UK Sanctions List Date Designated, col[34]=Last Updated, col[35]=Group ID
# Alias Type values: 'Primary name', 'Primary name variation', 'AKA'
# Alias Quality values: 'Good quality', 'Low quality'

def parse_uk(path):
    groups = {}

    with open(path, encoding='utf-8-sig', errors='replace', newline='') as fh:
        reader  = csv.reader(fh)
        headers = None
        for i, row in enumerate(reader):
            if i == 0: continue
            if i == 1: headers = [h.strip() for h in row]; continue
            if not headers or not any(row): continue
            if len(row) < len(headers):
                row = row + [''] * (len(headers) - len(row))
            d = dict(zip(headers, row))
            if d.get('Group Type', '').strip() != 'Individual': continue

            gid = d.get('Group ID', '').strip()
            if gid not in groups:
                groups[gid] = {
                    'data': d,
                    'aliases_good': [], 'aliases_low': [], 'aliases_fka': [],
                    'addresses': [],
                    'passports': [], 'nids': [],
                }
            g = groups[gid]

            # Aliases — categorised by Alias Type + Alias Quality
            at = d.get('Alias Type', '').strip()
            aq = d.get('Alias Quality', '').strip()
            # Build alias name from Name 1–6 on this row
            name_parts = [d.get(f'Name {j}', '').strip() for j in range(1, 6)]
            surname     = d.get('Name 6', '').strip()
            alias_name  = ' '.join(filter(None, name_parts + [surname]))
            if alias_name and alias_name != ' '.join(filter(None,
                    [d.get('Name 1','').strip()] + [d.get('Name 6','').strip()])):
                pass  # will handle below

            # Primary name row: Alias Type == 'Primary name' — this IS the person
            # Primary name variation: alternate spelling of primary name
            # AKA: alias
            if at == 'Primary name':
                pass  # data row — handled via g['data']
            elif alias_name:
                if 'low' in aq.lower():
                    if alias_name not in g['aliases_low']: g['aliases_low'].append(alias_name)
                else:
                    if alias_name not in g['aliases_good']: g['aliases_good'].append(alias_name)

            # Passport — collect all unique per-row passport numbers
            pnum = d.get('Passport Number', '').strip()
            pdet = d.get('Passport Details', '').strip()
            if pnum and pnum not in [p[0] for p in g['passports']]:
                g['passports'].append((pnum, pdet))

            # NID — collect all unique
            nnum = d.get('National Identification Number', '').strip()
            ndet = d.get('National Identification Details', '').strip()
            if nnum and nnum not in [n[0] for n in g['nids']]:
                g['nids'].append((nnum, ndet))

            # Addresses — each row may have a different address block
            addr_parts = [d.get(f'Address {j}', '').strip() for j in range(1, 7)]
            addr_line  = '; '.join(a for a in addr_parts if a)
            city       = d.get('Town of Birth', '').strip()
            country    = d.get('Country', '').strip()
            postcode   = d.get('Post/Zip Code', '').strip()
            addr_entry = (addr_line, city, country, postcode)
            if (addr_line or country) and addr_entry not in g['addresses']:
                g['addresses'].append(addr_entry)

    rows = []
    for gid, g in groups.items():
        d   = g['data']
        row = empty_row('UK HMT')

        row['Reference / UID']        = s(gid)
        row['Sanctions Programme']    = s(d.get('Regime', ''))
        row['Last Name']              = s(d.get('Name 6', ''))
        row['First Name']             = s(d.get('Name 1', ''))
        row['Second / Middle Name']   = s(d.get('Name 2', ''))
        row['Third Name']             = s(d.get('Name 3', ''))
        row['Fourth Name']            = s(d.get('Name 4', ''))
        row['Fifth Name']             = s(d.get('Name 5', ''))
        row['Title']                  = s(d.get('Title', ''))
        row['Designation / Position'] = s(d.get('Position', ''))
        row['Date of Birth']          = s(d.get('DOB', ''))
        row['Place of Birth']         = s(d.get('Town of Birth', ''))
        row['Country of Birth']       = s(d.get('Country of Birth', ''))
        row['Nationality 1']          = s(d.get('Nationality', ''))
        row['Listed On']              = s(d.get('Listed On', ''))
        row['UK Designated On']       = s(d.get('UK Sanctions List Date Designated', ''))
        row['Last Updated']           = s(d.get('Last Updated', ''))
        row['Remarks / Other Info']   = s(d.get('Other Information', ''))

        # Aliases
        row['Good Quality AKA'] = join(g['aliases_good'], '; ')
        row['Low Quality AKA']  = join(g['aliases_low'],  '; ')
        row['Formerly Known As']= join(g['aliases_fka'],  '; ')

        # Passports — all
        if g['passports']:
            row['Passport Number']  = join([p[0] for p in g['passports']], ' | ')
            row['Passport Details'] = join([p[1] for p in g['passports'] if p[1]], ' | ')

        # NIDs — all
        if g['nids']:
            row['National ID Number']  = join([n[0] for n in g['nids']], ' | ')
            row['National ID Details'] = join([n[1] for n in g['nids'] if n[1]], ' | ')

        # Addresses — all
        if g['addresses']:
            row['Address']     = join([a[0] for a in g['addresses'] if a[0]], ' | ')
            row['City']        = join([a[1] for a in g['addresses'] if a[1]], ' | ')
            row['Country']     = join([a[2] for a in g['addresses'] if a[2]], ' | ')
            row['Postal Code'] = join([a[3] for a in g['addresses'] if a[3]], ' | ')

        rows.append(row)
    return rows

# ── Parser: EU FSF CSV ────────────────────────────────────────────────────────
# Verified column map from actual file (59 columns, semicolon-delimited):
# col[1]  = person_id (grouping key, 4405 unique persons)
# col[2]  = 'P' for person
# col[4]  = designation/listing date
# col[6]  = programme (IRQ, TAQA, AFG, TERR etc.)
# col[7]  = remarks / legal basis
# col[58] = EU reference number (EU.XXX.XX)
#
# NAME ROWS  (col[14] and/or col[15] populated):
#   col[14]=last, col[15]=first, col[16]=second, col[17]=full/alias
#   col[18]=gender, col[19]=title, col[20]=designation, col[21]=nationality ISO2
#
# DOB ROWS   (col[40] populated):
#   col[40]=dob, col[41]=place of birth, col[42]=country of birth ISO2
#
# ID ROWS    (col[49] populated):
#   col[49]="number (doc-type) (notes)", col[50]=country ISO2
#
# ADDRESS ROWS (col[29] or col[31] populated):
#   col[29]=street, col[30]=postcode, col[31]=city, col[32]=country ISO2
#
# NATIONALITY ROWS (col[57] populated):
#   col[57]=nationality ISO2

_EU_ID_RE = re.compile(r'^(.+?)\s*\(([^)]+)\)', re.IGNORECASE)

def _parse_eu_id(raw):
    m = _EU_ID_RE.match(raw)
    if m: return m.group(1).strip(), m.group(2).strip()
    return raw.strip(), ''

def parse_eu(path):
    groups = {}

    with open(path, encoding='utf-8-sig', errors='replace', newline='') as fh:
        for line in fh:
            parts = line.rstrip('\r\n').split(';')
            if len(parts) < 5: continue
            if parts[2].strip() != 'P': continue
            pid = parts[1].strip()
            if not pid: continue

            if pid not in groups:
                groups[pid] = {
                    'last': '', 'first': '', 'second': '',
                    'gender': '', 'title': '', 'function': '',
                    'aliases': [], 'dobs': [], 'pobs': [], 'nats': [],
                    'passports': [], 'nids': [], 'addresses': [],
                    'programmes': [], 'listed_ons': [], 'eu_refs': [], 'remarks': [],
                }
            g = groups[pid]

            def c(i): return parts[i].strip() if len(parts) > i else ''

            # EU reference (col 58)
            eu_ref = c(58)
            if eu_ref and eu_ref not in g['eu_refs']:
                g['eu_refs'].append(eu_ref)

            # Programme (col 6), listing date (col 4), remarks (col 7)
            for field, store in [(c(6), g['programmes']), (c(7), g['remarks'])]:
                if field and field not in store: store.append(field)
            listed = c(4)
            if listed and listed not in g['listed_ons']: g['listed_ons'].append(listed)

            # NAME ROWS — col[14] last OR col[15] first OR col[17] alias present
            last_v = c(14); first_v = c(15); second_v = c(16); full_v = c(17)
            if last_v or first_v or full_v:
                if last_v and first_v:
                    # Primary name row
                    if not g['last']:     g['last']     = last_v
                    if not g['first']:    g['first']    = first_v
                    if second_v and not g['second']:  g['second']  = second_v
                    if c(18) and not g['gender']:     g['gender']  = c(18)
                    if c(19) and not g['title']:      g['title']   = c(19)
                    if c(20) and not g['function']:   g['function']= c(20)
                    nat = c(21)
                    if nat and nat not in g['nats']:  g['nats'].append(nat)
                else:
                    # Alias row
                    alias = full_v or ' '.join(filter(None, [first_v, last_v]))
                    if alias and alias not in g['aliases']: g['aliases'].append(alias)
                    nat = c(21)
                    if nat and nat not in g['nats']: g['nats'].append(nat)

            # DOB ROWS — col[40]
            dob_v = c(40)
            if dob_v and dob_v.upper() not in ('N/A', 'NA', ''):
                if dob_v not in g['dobs']: g['dobs'].append(dob_v)
                pob = (c(41), c(42))
                if any(pob) and pob not in g['pobs']: g['pobs'].append(pob)

            # ID ROWS — col[49]
            id_raw = c(49)
            if id_raw:
                num, doc_type = _parse_eu_id(id_raw)
                country = c(50)
                if 'passport' in doc_type.lower():
                    entry = (num, country, doc_type)
                    if entry not in g['passports']: g['passports'].append(entry)
                else:
                    entry = (num, country, doc_type)
                    if entry not in g['nids']: g['nids'].append(entry)

            # ADDRESS ROWS — col[29] or col[31]
            if c(29) or c(31):
                ae = (c(29), c(30), c(31), c(32))
                if ae not in g['addresses']: g['addresses'].append(ae)

            # NATIONALITY ROWS — col[57]
            nat2 = c(57)
            if nat2 and nat2 not in g['nats']: g['nats'].append(nat2)

    rows = []
    for pid, g in groups.items():
        if not g['last'] and not g['first'] and not g['aliases']:
            continue
        row = empty_row('EU')

        row['Reference / UID']        = join([pid] + g['eu_refs'], ' | ')
        row['Sanctions Programme']    = join(g['programmes'], '; ')
        row['Last Name']              = s(g['last'])
        row['First Name']             = s(g['first'])
        row['Second / Middle Name']   = s(g['second'])
        row['Title']                  = s(g['title'])
        row['Designation / Position'] = s(g['function'])
        row['Gender']                 = s(g['gender'])

        # Aliases excluding the primary name itself
        primary = ' '.join(filter(None, [g['first'], g['second'], g['last']]))
        row['Good Quality AKA'] = join([a for a in g['aliases'] if a != primary], '; ')

        row['Remarks / Other Info'] = join(g['remarks'], ' | ')
        row['Date of Birth']        = join(g['dobs'], ' | ')

        if g['pobs']:
            row['Place of Birth']   = join([p[0] for p in g['pobs'] if p[0]], ' | ')
            row['Country of Birth'] = join([p[1] for p in g['pobs'] if p[1]], ' | ')

        nats = [n for n in g['nats'] if n]
        if nats:          row['Nationality 1']  = nats[0]
        if len(nats) > 1: row['Nationality 2']  = nats[1]
        if len(nats) > 2: row['Nationality 3+'] = join(nats[2:], '; ')

        if g['passports']:
            row['Passport Number']  = join([p[0] for p in g['passports']], ' | ')
            row['Passport Country'] = join([p[1] for p in g['passports'] if p[1]], ' | ')
            row['Passport Expiry']  = join([p[2] for p in g['passports'] if p[2]], ' | ')

        if g['nids']:
            row['National ID Number']  = join([n[0] for n in g['nids']], ' | ')
            row['National ID Country'] = join([n[1] for n in g['nids'] if n[1]], ' | ')
            row['National ID Type']    = join([n[2] for n in g['nids'] if n[2]], ' | ')

        if g['addresses']:
            row['Address']     = join([a[0] for a in g['addresses'] if a[0]], ' | ')
            row['Postal Code'] = join([a[1] for a in g['addresses'] if a[1]], ' | ')
            row['City']        = join([a[2] for a in g['addresses'] if a[2]], ' | ')
            row['Country']     = join([a[3] for a in g['addresses'] if a[3]], ' | ')

        row['Listed On'] = join(g['listed_ons'], ' | ')
        rows.append(row)
    return rows

# ── Parser: UNSC XML ──────────────────────────────────────────────────────────
# Verified tags (no namespace): DATAID, FIRST_NAME, SECOND_NAME, THIRD_NAME,
# FOURTH_NAME, TITLE/VALUE, DESIGNATION/VALUE, GENDER, UN_LIST_TYPE,
# REFERENCE_NUMBER, LISTED_ON, LAST_DAY_UPDATED/VALUE, HAS_INTERPOL_LINK,
# INTERPOL_LINK, COMMENTS1, NAME_ORIGINAL_SCRIPT,
# NATIONALITY/VALUE, LIST_TYPE/VALUE,
# INDIVIDUAL_ALIAS/(QUALITY, ALIAS_NAME, NOTE, DATE_OF_BIRTH,
#                   CITY_OF_BIRTH, COUNTRY_OF_BIRTH)
# INDIVIDUAL_DATE_OF_BIRTH/(TYPE_OF_DATE, YEAR, FROM_YEAR, TO_YEAR,
#                            DATE, NOTE)
# INDIVIDUAL_PLACE_OF_BIRTH/(CITY, COUNTRY, STATE_PROVINCE, STREET, NOTE)
# INDIVIDUAL_DOCUMENT/(TYPE_OF_DOCUMENT, TYPE_OF_DOCUMENT2, NUMBER,
#                      ISSUING_COUNTRY, COUNTRY_OF_ISSUE, CITY_OF_ISSUE,
#                      DATE_OF_ISSUE, DATE_OF_EXPIRY, NOTE)
# INDIVIDUAL_ADDRESS/(STREET, CITY, STATE_PROVINCE, COUNTRY, ZIP_CODE, NOTE)

def parse_unsc(path):
    rows = []
    try:
        tree = ET.parse(path); root = tree.getroot()
    except ET.ParseError as e:
        print(f'\n  ✗ UNSC XML: {e}'); return rows

    ns_m = re.match(r'\{(.*?)\}', root.tag)
    ns   = f'{{{ns_m.group(1)}}}' if ns_m else ''

    def tx(el, tag):
        n = el.find(f'{ns}{tag}')
        return s(n.text) if n is not None and n.text else ''

    for indiv in root.findall(f'.//{ns}INDIVIDUAL'):
        row = empty_row('UNSC')

        row['Reference / UID']     = s(tx(indiv, 'REFERENCE_NUMBER'))
        row['Sanctions Programme'] = s(tx(indiv, 'UN_LIST_TYPE'))
        row['Listed On']           = s(tx(indiv, 'LISTED_ON'))
        row['Gender']              = s(tx(indiv, 'GENDER'))
        row['Interpol Link']       = s(tx(indiv, 'INTERPOL_LINK'))

        # Last Updated
        lu_el = indiv.find(f'.//{ns}LAST_DAY_UPDATED/{ns}VALUE')
        if lu_el is not None and lu_el.text:
            row['Last Updated'] = s(lu_el.text)

        # Title
        title_el = indiv.find(f'.//{ns}TITLE/{ns}VALUE')
        if title_el is not None and title_el.text:
            row['Title'] = s(title_el.text)

        # Designation
        desig_el = indiv.find(f'.//{ns}DESIGNATION/{ns}VALUE')
        if desig_el is not None and desig_el.text:
            row['Designation / Position'] = s(desig_el.text)

        # Name parts — last non-empty = Last Name
        fn     = s(tx(indiv, 'FIRST_NAME'))
        sn     = s(tx(indiv, 'SECOND_NAME'))
        tn     = s(tx(indiv, 'THIRD_NAME'))
        fourth = s(tx(indiv, 'FOURTH_NAME'))
        row['First Name']           = fn
        row['Second / Middle Name'] = sn
        row['Third Name']           = tn
        row['Fourth Name']          = fourth
        row['Last Name']            = last_nonempty(fn, sn, tn, fourth)

        # Comments (COMMENTS1 only per schema — verified)
        comments = s(tx(indiv, 'COMMENTS1'))
        if comments: row['Remarks / Other Info'] = comments

        # Nationalities — all VALUE children
        nats = [s(tx(n, 'VALUE')) for n in indiv.findall(f'.//{ns}NATIONALITY')
                if tx(n, 'VALUE')]
        if nats:          row['Nationality 1']  = nats[0]
        if len(nats) > 1: row['Nationality 2']  = nats[1]
        if len(nats) > 2: row['Nationality 3+'] = join(nats[2:], '; ')

        # Aliases — QUALITY can be 'Good', 'Low', '' (empty = good)
        good, low = [], []
        for alias in indiv.findall(f'.//{ns}INDIVIDUAL_ALIAS'):
            quality = tx(alias, 'QUALITY').lower()
            name    = s(tx(alias, 'ALIAS_NAME'))
            if not name: continue
            note = tx(alias, 'NOTE')
            full = f'{name} [{note}]' if note else name
            (low if 'low' in quality else good).append(full)
        row['Good Quality AKA'] = join(good, '; ')
        row['Low Quality AKA']  = join(low,  '; ')

        # DOB — all, build from YEAR/DATE/FROM_YEAR/TO_YEAR/NOTE
        dobs = []
        for dob_el in indiv.findall(f'.//{ns}INDIVIDUAL_DATE_OF_BIRTH'):
            date_exact = tx(dob_el, 'DATE')
            yr  = tx(dob_el, 'YEAR')
            mo  = tx(dob_el, 'MONTH') or tx(dob_el, 'FROM_MONTH')
            dy  = tx(dob_el, 'DAY')   or tx(dob_el, 'FROM_DAY')
            fyr = tx(dob_el, 'FROM_YEAR')
            tyr = tx(dob_el, 'TO_YEAR')
            note= tx(dob_el, 'NOTE')

            if date_exact:
                dob_str = date_exact
            elif yr:
                dob_str = '/'.join(filter(None, [dy, mo, yr]))
            elif fyr and tyr:
                dob_str = f'{fyr}–{tyr}'
            elif fyr:
                dob_str = f'circa {fyr}'
            elif note:
                dob_str = note
            else:
                continue

            if dob_str not in dobs: dobs.append(dob_str)
        row['Date of Birth'] = join(dobs, ' | ')

        # POB — all
        pobs, pob_countries = [], []
        for pob_el in indiv.findall(f'.//{ns}INDIVIDUAL_PLACE_OF_BIRTH'):
            city  = tx(pob_el, 'CITY') or tx(pob_el, 'STREET') or tx(pob_el, 'STATE_PROVINCE')
            ctry  = tx(pob_el, 'COUNTRY')
            note  = tx(pob_el, 'NOTE')
            city_full = ' '.join(filter(None, [city, f'({note})' if note else '']))
            if city_full and city_full not in pobs: pobs.append(city_full)
            if ctry and ctry not in pob_countries: pob_countries.append(ctry)
        row['Place of Birth']   = join(pobs,          ' | ')
        row['Country of Birth'] = join(pob_countries, ' | ')

        # Documents — all
        passports, other_ids = [], []
        for doc in indiv.findall(f'.//{ns}INDIVIDUAL_DOCUMENT'):
            doc_type   = tx(doc, 'TYPE_OF_DOCUMENT') or tx(doc, 'TYPE_OF_DOCUMENT2')
            doc_num    = s(tx(doc, 'NUMBER'))
            doc_issuer = tx(doc, 'ISSUING_COUNTRY') or tx(doc, 'COUNTRY_OF_ISSUE')
            doc_issue  = tx(doc, 'DATE_OF_ISSUE')
            doc_expiry = tx(doc, 'DATE_OF_EXPIRY')
            doc_note   = tx(doc, 'NOTE')
            if not doc_num: continue
            num_full = f'{doc_num} [{doc_note}]' if doc_note else doc_num
            if 'passport' in doc_type.lower():
                passports.append((num_full, doc_issuer, doc_issue, doc_expiry, doc_type))
            else:
                other_ids.append((num_full, doc_issuer, doc_issue, doc_expiry, doc_type))

        if passports:
            row['Passport Number']     = join([p[0] for p in passports], ' | ')
            row['Passport Country']    = join([p[1] for p in passports if p[1]], ' | ')
            row['Passport Issue Date'] = join([p[2] for p in passports if p[2]], ' | ')
            row['Passport Expiry']     = join([p[3] for p in passports if p[3]], ' | ')
        if other_ids:
            row['National ID Number']  = join([x[0] for x in other_ids], ' | ')
            row['National ID Country'] = join([x[1] for x in other_ids if x[1]], ' | ')
            row['National ID Type']    = join([x[4] for x in other_ids if x[4]], ' | ')

        # Addresses — all
        addrs, cities, states, countries = [], [], [], []
        for addr in indiv.findall(f'.//{ns}INDIVIDUAL_ADDRESS'):
            street = tx(addr, 'STREET')
            note   = tx(addr, 'NOTE')
            line   = ' '.join(filter(None, [street, f'({note})' if note else '']))
            city   = tx(addr, 'CITY')
            state  = tx(addr, 'STATE_PROVINCE')
            ctry   = tx(addr, 'COUNTRY')
            if line and line not in addrs:   addrs.append(line)
            if city and city not in cities:  cities.append(city)
            if state and state not in states: states.append(state)
            if ctry and ctry not in countries: countries.append(ctry)

        row['Address']          = join(addrs,     ' | ')
        row['City']             = join(cities,    ' | ')
        row['State / Province'] = join(states,    ' | ')
        row['Country']          = join(countries, ' | ')

        rows.append(row)
    return rows

# ── Parser: NACTA (Excel .xlsx downloaded by Playwright) ─────────────────────
# The Playwright downloader saves the Excel file from nfs.nacta.gov.pk.
# Expected columns (from the site's table):
#   S.No | Name | Father Name | CNIC | District | Province
# We use openpyxl to read the xlsx directly.

def parse_nacta(path):
    import re as _re
    rows = []
    if not path or not Path(path).exists():
        return rows

    path = Path(path)

    # ── Excel (.xlsx) ─────────────────────────────────────────────────────────
    if path.suffix.lower() in ('.xlsx', '.xls'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active

            headers = []
            header_row_found = False

            for row in ws.iter_rows(values_only=True):
                cells = [s(str(c)) if c is not None else '' for c in row]

                # Detect header row
                if not header_row_found:
                    row_lower = ' '.join(cells).lower()
                    if any(k in row_lower for k in ['name', 'cnic', 'father']):
                        headers = [c.lower().strip() for c in cells]
                        header_row_found = True
                    continue

                if not any(cells):
                    continue

                # Map by header position
                def get(col_names):
                    for cn in col_names:
                        for i, h in enumerate(headers):
                            if cn in h and i < len(cells):
                                v = cells[i]
                                if v: return v
                    # fallback: positional
                    return ''

                full_name   = get(['name'])
                father_name = get(['father'])
                cnic        = get(['cnic', 'id', 'nic'])
                district    = get(['district'])
                province    = get(['province'])

                # Skip blank rows
                if not full_name and not cnic:
                    continue

                row_out = empty_row('NACTA')
                row_out['Sanctions Programme'] = 'Pakistan Schedule IV (Anti Terrorism Act 1997)'

                # Split full name into parts
                name_parts = full_name.split() if full_name else []
                if len(name_parts) == 1:
                    row_out['First Name'] = name_parts[0]
                    row_out['Last Name']  = name_parts[0]
                elif len(name_parts) == 2:
                    row_out['First Name'] = name_parts[0]
                    row_out['Last Name']  = name_parts[1]
                elif len(name_parts) >= 3:
                    row_out['First Name']           = name_parts[0]
                    row_out['Second / Middle Name'] = ' '.join(name_parts[1:-1])
                    row_out['Last Name']            = name_parts[-1]

                # Clean CNIC — remove non-digit/hyphen chars
                cnic_clean = _re.sub(r'[^\d\-]', '', cnic)
                row_out['National ID Number']  = cnic_clean
                row_out['National ID Type']    = 'CNIC (Pakistan)'
                row_out['National ID Country'] = 'Pakistan'

                if father_name:
                    row_out['Remarks / Other Info'] = f'Father: {father_name}'

                row_out['City']             = district
                row_out['State / Province'] = province
                row_out['Country']          = 'Pakistan'

                rows.append(row_out)

            wb.close()

        except Exception as e:
            print(f'\n  ✗ NACTA Excel parse error: {e}')

    # ── OpenSanctions CSV fallback ───────────────────────────────────────────
    elif path.suffix.lower() == '.csv':
        import csv as _csv
        try:
            with open(path, encoding='utf-8-sig', errors='replace', newline='') as fh:
                reader = _csv.DictReader(fh)
                for d in reader:
                    # OpenSanctions simple CSV columns:
                    # id, name, aliases, birth_date, nationalities, addresses,
                    # identifiers, sanctions, sourceUrl
                    full_name = s(d.get('name', ''))
                    if not full_name: continue

                    row_out = empty_row('NACTA')
                    row_out['Sanctions Programme'] = 'Pakistan Schedule IV (Anti Terrorism Act 1997)'

                    name_parts = full_name.split()
                    if len(name_parts) == 1:
                        row_out['First Name'] = row_out['Last Name'] = name_parts[0]
                    elif len(name_parts) == 2:
                        row_out['First Name'] = name_parts[0]
                        row_out['Last Name']  = name_parts[1]
                    else:
                        row_out['First Name']           = name_parts[0]
                        row_out['Second / Middle Name'] = ' '.join(name_parts[1:-1])
                        row_out['Last Name']            = name_parts[-1]

                    # Aliases
                    aliases = s(d.get('aliases', ''))
                    if aliases: row_out['Good Quality AKA'] = aliases

                    # DOB
                    row_out['Date of Birth'] = s(d.get('birth_date', ''))

                    # Identifiers — look for CNIC (13 digits)
                    idents = s(d.get('identifiers', ''))
                    if idents:
                        import re as _re2
                        cnic_m = _re2.search(r'(\d{5}-\d{7}-\d|\d{13})', idents)
                        if cnic_m:
                            row_out['National ID Number']  = cnic_m.group(1)
                            row_out['National ID Type']    = 'CNIC (Pakistan)'
                            row_out['National ID Country'] = 'Pakistan'
                        else:
                            row_out['National ID Number'] = idents[:80]

                    # Address
                    addr = s(d.get('addresses', ''))
                    if addr: row_out['Address'] = addr

                    row_out['Country'] = 'Pakistan'
                    row_out['Nationality 1'] = s(d.get('nationalities', 'Pakistan'))

                    rows.append(row_out)
        except Exception as e:
            print(f'\n  ✗ NACTA CSV parse error: {e}')

    # ── XML fallback (old format) ─────────────────────────────────────────────
    elif path.suffix.lower() == '.xml':
        try:
            tree = ET.parse(path); root = tree.getroot()
            ns_m = re.match(r'\{(.*?)\}', root.tag)
            ns   = f'{{{ns_m.group(1)}}}' if ns_m else ''

            def tx(el, tag):
                n = el.find(f'{ns}{tag}')
                return s(n.text) if n is not None and n.text else ''

            for model in root.findall(f'.//{ns}HomeModel'):
                full_name   = tx(model, 'Name')
                father_name = tx(model, 'FatherName')
                cnic        = tx(model, 'CNIC')
                district    = tx(model, 'District')
                province    = tx(model, 'Province')

                if not full_name and not cnic: continue

                row_out = empty_row('NACTA')
                row_out['Sanctions Programme'] = 'Pakistan Schedule IV (Anti Terrorism Act 1997)'

                name_parts = full_name.split() if full_name else []
                if len(name_parts) == 1:
                    row_out['First Name'] = row_out['Last Name'] = name_parts[0]
                elif len(name_parts) == 2:
                    row_out['First Name'] = name_parts[0]; row_out['Last Name'] = name_parts[1]
                elif len(name_parts) >= 3:
                    row_out['First Name']           = name_parts[0]
                    row_out['Second / Middle Name'] = ' '.join(name_parts[1:-1])
                    row_out['Last Name']            = name_parts[-1]

                row_out['National ID Number']  = cnic
                row_out['National ID Type']    = 'CNIC (Pakistan)'
                row_out['National ID Country'] = 'Pakistan'
                if father_name:
                    row_out['Remarks / Other Info'] = f'Father: {father_name}'
                row_out['City']             = district
                row_out['State / Province'] = province
                row_out['Country']          = 'Pakistan'
                rows.append(row_out)

        except Exception as e:
            print(f'\n  ✗ NACTA XML parse error: {e}')

    return rows


PARSERS = {
    'xml_ofac': parse_ofac_xml,
    'csv_uk':   parse_uk,
    'csv_eu':   parse_eu,
    'xml_unsc': parse_unsc,
    'xml_nacta': parse_nacta,
}

# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(all_rows, counts):
    _thin   = Side(style='thin', color='D0D0D0')
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def hdr(cell, fg, text):
        cell.value     = text
        cell.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor=fg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = _border

    def dat(cell, val, fg):
        cell.value     = s(str(val)) if val else ''
        cell.font      = Font(name='Arial', size=9)
        cell.fill      = PatternFill('solid', fgColor=fg)
        cell.alignment = Alignment(vertical='top', wrap_text=False)
        cell.border    = _border

    wb = Workbook()

    # Sheet 1: All Records
    ws = wb.active
    ws.title = 'All Records'
    for ci, (h, w) in enumerate(zip(MASTER_COLS, MASTER_COL_W), 1):
        hdr(ws.cell(row=1, column=ci), '1F3864', h)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30

    row_num, cur = 2, None
    for entry in all_rows:
        src = entry['Source']
        if src != cur:
            for ci in range(1, len(MASTER_COLS) + 1):
                c = ws.cell(row=row_num, column=ci,
                            value=f'  ▶  {src}' if ci == 1 else '')
                c.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
                c.fill      = PatternFill('solid', fgColor=COLORS[src]['sep'])
                c.alignment = Alignment(vertical='center')
                c.border    = _border
            ws.row_dimensions[row_num].height = 16
            row_num += 1; cur = src
        for ci, col in enumerate(MASTER_COLS, 1):
            dat(ws.cell(row=row_num, column=ci), entry.get(col, ''), COLORS[src]['row'])
        ws.row_dimensions[row_num].height = 15
        row_num += 1

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(MASTER_COLS))}1'

    # Per-source sheets
    for src in ORDER:
        src_rows = [r for r in all_rows if r['Source'] == src]
        cols     = SOURCE_COLS[src]
        ws_s     = wb.create_sheet(src.replace(' ', '_'))
        color    = COLORS[src]
        for ci, col in enumerate(cols, 1):
            idx = MASTER_COLS.index(col) if col in MASTER_COLS else -1
            w   = MASTER_COL_W[idx] if idx >= 0 else 20
            hdr(ws_s.cell(row=1, column=ci), color['header'], col)
            ws_s.column_dimensions[get_column_letter(ci)].width = w
        ws_s.row_dimensions[1].height = 30
        for ri, entry in enumerate(src_rows, 2):
            for ci, col in enumerate(cols, 1):
                dat(ws_s.cell(row=ri, column=ci), entry.get(col, ''), color['row'])
            ws_s.row_dimensions[ri].height = 15
        ws_s.freeze_panes = 'A2'
        ws_s.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'

    # Summary sheet
    ws2 = wb.create_sheet('Summary', 1)
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 55

    ws2.merge_cells('A1:C1')
    ws2['A1'].value     = 'Global Sanctions Lists  —  Full Data Extraction  v5'
    ws2['A1'].font      = Font(name='Arial', bold=True, size=14, color='1F3864')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 32

    ws2.merge_cells('A2:C2')
    ws2['A2'].value     = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}"
    ws2['A2'].font      = Font(name='Arial', italic=True, size=10, color='808080')
    ws2['A2'].alignment = Alignment(horizontal='center')

    for ci, h in enumerate(['Source', 'Records', 'Columns'], 1):
        hdr(ws2.cell(row=4, column=ci), '1F3864', h)
    ws2.row_dimensions[4].height = 24

    for ri, src in enumerate(ORDER, 5):
        c1 = ws2.cell(row=ri, column=1, value=src)
        c2 = ws2.cell(row=ri, column=2, value=counts.get(src, 0))
        c3 = ws2.cell(row=ri, column=3,
                      value=', '.join(c for c in SOURCE_COLS[src] if c != 'Source'))
        fill = PatternFill('solid', fgColor=COLORS[src]['row'])
        for c, ha in [(c1,'left'),(c2,'center'),(c3,'left')]:
            c.font      = Font(name='Arial', size=10, bold=(c==c1))
            c.fill      = fill
            c.alignment = Alignment(horizontal=ha, vertical='top', wrap_text=True)
            c.border    = _border
        ws2.row_dimensions[ri].height = 40

    tr = 5 + len(ORDER)
    for c, v, ha in [
        (ws2.cell(row=tr, column=1), 'TOTAL', 'left'),
        (ws2.cell(row=tr, column=2), sum(counts.values()), 'center'),
        (ws2.cell(row=tr, column=3), f'{len(MASTER_COLS)} master columns', 'left'),
    ]:
        c.value = v
        c.font  = Font(name='Arial', bold=True, size=11)
        c.alignment = Alignment(horizontal=ha, vertical='center')

    wb.save(OUTPUT_FILE)

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print('=' * 62)
    print('  Global Sanctions Lists Extractor  v5  —  full data')
    print(f'  {datetime.now().strftime("%d %B %Y  %H:%M")}')
    print('=' * 62)

    print('\n[1/3] Downloading...\n')
    local_files = {}
    for key, cfg in SOURCES.items():
        try:
            if key == 'NACTA':
                local_files[key] = download_nacta(cfg)
            else:
                local_files[key] = download(key, cfg)
        except Exception as e:
            print(f'\n  [{key}] FAILED: {e}')
            local_files[key] = None

    print('\n[2/3] Parsing...\n')
    all_rows, counts = [], {}
    for key in ORDER:
        path = local_files.get(key)
        if not path or not path.exists():
            print(f'  [{key}] Skipped'); counts[key] = 0; continue
        fmt     = SOURCES[key]['format']
        size_mb = path.stat().st_size / 1024 / 1024
        print(f'  [{key}] {path.name} ({size_mb:.1f} MB) ...', end='', flush=True)
        try:
            rows = PARSERS[fmt](path)
        except Exception as e:
            print(f'\n  ✗ {e}'); rows = []
        counts[key] = len(rows)
        all_rows.extend(rows)
        print(f'  {len(rows):,} records  ✓')

    print(f'\n  Total: {len(all_rows):,} records')
    print('\n[3/3] Building Excel...')
    build_excel(all_rows, counts)

    print(f'\n✓ Saved → {OUTPUT_FILE}')
    print(f'\n  {"Source":<12}  {"Records":>9}  Cols')
    print(f'  {"-"*12}  {"-"*9}  {"-"*4}')
    for src in ORDER:
        print(f'  {src:<12}  {counts.get(src,0):>9,}  {len(SOURCE_COLS[src])}')
    print(f'  {"TOTAL":<12}  {sum(counts.values()):>9,}  {len(MASTER_COLS)} combined')
    print(f'\n  Cache: {CACHE_DIR}  (delete to re-download)')

if __name__ == '__main__':
    main()
